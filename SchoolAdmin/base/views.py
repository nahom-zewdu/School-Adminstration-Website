from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render, HttpResponse, redirect, HttpResponsePermanentRedirect, get_object_or_404
from django.urls import reverse
from django.contrib import messages
from django.contrib.auth.models import User
from .models import *
from .forms import *
from .utils import *
from openpyxl import load_workbook

# Create your views here.

# base views section
def home(request):
    students = Student.objects.all()
    teachers = Teacher.objects.all()
    staffs = Staff.objects.all()

    female_students = Student.filter_by_gender('F')
    male_students = Student.filter_by_gender('M')

    female_teachers = Teacher.filter_by_gender('F')
    male_teachers = Teacher.filter_by_gender('M')
    
    female_staffs = Staff.filter_by_gender('F')
    male_staffs = Staff.filter_by_gender('M')
    
    context = {
        'teachers': teachers,
        'students': students,
        'staffs': staffs,

        'female_students': female_students,
        'male_students': male_students,

        'female_teachers': female_teachers,
        'male_teachers': male_teachers,

        'female_staffs': female_staffs,
        'male_staffs': male_staffs,
    }

    return render(request, 'base/home.html', context)


@login_required(login_url='/restricted/')
def student_academics(request):
    students = Student.objects.all()
    female = Student.filter_by_gender('F')
    male = Student.filter_by_gender('M')
    context = {
        'students': students,
        'female': female,
        'male': male,
    }

    return render(request, 'student_academics.html', context)

@login_required(login_url='/restricted/')
def teacher_academics(request):
    teachers = Teacher.objects.all()
    female = Teacher.filter_by_gender('F')
    male = Teacher.filter_by_gender('M')
    context = {
        'teachers': teachers,
        'female': female,
        'male': male,
    }
    return render(request, 'teacher_academics.html', context)



# Authentication views section
def logout_view(request):
    logout(request)

    return HttpResponsePermanentRedirect(reverse('base:home'))

def student_login(request):
    if request.method == 'POST':
        student_id = request.POST['student_id']
        password = request.POST['password']

        user = authenticate(request, student_id=student_id, password=password)
        if user:
            login(request, user)
            return redirect('base:home')
        else:
            error_message = 'Invalid ID number or Password'
            return render(request, 'login/student_login.html', {'error_message': error_message})
    else:
        return render(request, 'login/student_login.html')

def teacher_login(request):
    if request.method == 'POST':
        teacher_id = request.POST['teacher_id']
        password = request.POST['password']

        user = authenticate(request, teacher_id=teacher_id, password=password)
        if user:
            login(request, user)
            return redirect('base:home')
        else:
            error_message = 'Invalid ID number or Password'
            return render(request, 'login/teacher_login.html', {'error_message': error_message})
    else:
        return render(request, 'login/teacher_login.html')

def staff_login(request):
    if request.method == 'POST':
        staff_id = request.POST['staff_id']
        password = request.POST['password']

        user = authenticate(request, staff_id=staff_id, password=password)
        if user:
            login(request, user)
            if user.is_staff:
                return redirect('base:dashboard')
            else:
                return redirect('base:home')
        else:
            error_message = 'Invalid ID number or Password'
            return render(request, 'login/staff_login.html', {'error_message': error_message})
    else:
        return render(request, 'login/staff_login.html')

def parent_login(request):
    if request.method == 'POST':
        parent_id = request.POST['parent_id']
        password = request.POST['password']

        user = authenticate(request, parent_id=parent_id, password=password)
        if user:
            login(request, user)
            return redirect('base:home')
        else:
            error_message = 'Invalid ID number or Password'
            return render(request, 'login/parent_login.html', {'error_message': error_message})
    else:
        return render(request, 'login/parent_login.html')



# Dashboard Views section
@login_required
@user_passes_test(lambda user: user.is_staff)
def dashboard(request):
    students = Student.objects.all()
    teachers = Teacher.objects.all()
    staffs = Staff.objects.all()
    parents = Parent.objects.all()

    female_students = Student.filter_by_gender('F')
    male_students = Student.filter_by_gender('M')

    female_teachers = Teacher.filter_by_gender('F')
    male_teachers = Teacher.filter_by_gender('M')
    
    female_staffs = Staff.filter_by_gender('F')
    male_staffs = Staff.filter_by_gender('M')
    
    context = {
        'teachers': teachers,
        'students': students,
        'staffs': staffs,
        'parents': parents,

        'female_students': female_students,
        'male_students': male_students,

        'female_teachers': female_teachers,
        'male_teachers': male_teachers,

        'female_staffs': female_staffs,
        'male_staffs': male_staffs,
    }
    return render(request, 'dashboard/dashboard.html', context)


@login_required
@user_passes_test(lambda user: user.is_staff)
def teacher_register(request):
    if request.method == 'POST':
        form = TeacherCreationForm(request.POST)
        if form.is_valid():
            first_name = form.cleaned_data.get('first_name')
            last_name = form.cleaned_data.get('last_name')
            email = form.cleaned_data.get('email')
            username = first_name.strip().lower() + '_' + last_name.strip().lower()

            if User.objects.filter(username=username).exists():
                messages.error(request, f'{username} is already registered. Please choose a different username.')
            else:
                user = User.objects.create_user(
                    username=username, 
                    first_name=first_name, 
                    last_name=last_name,
                    email=email,
                    password='@@aa12345',
                )
                if user:
                    name = first_name + ' ' + last_name
                    gender = form.cleaned_data.get('gender')
                    department = form.cleaned_data.get('department')
                    experience = form.cleaned_data.get('experience')
                    phone = form.cleaned_data.get('phone')
                    teacher = Teacher(
                        user=user, 
                        name=name, 
                        gender=gender,
                        department=department, 
                        experience=experience,
                        phone=phone,
                    )
                    if teacher:
                        teacher.save()
                        messages.success(request, f'{name} is registered successfully!')
                        return redirect('base:dashboard')
                    else:
                        user.delete()
                        messages.error(request, 'Someting went wrong while registering Teacher')

                return redirect('base:teacher_register')
            
    else:
        form = TeacherCreationForm()

    context = {
        'form': form,
        'type': 'teacher'
    }
    return render(request, 'dashboard/register.html', context)


@login_required
@user_passes_test(lambda user: user.is_staff)
def student_register(request):
    if request.method == 'POST':
        form = StudentCreationForm(request.POST, request.FILES)
        if form.is_valid():
            print(form)
            first_name = form.cleaned_data.get('first_name')
            last_name = form.cleaned_data.get('last_name')
            email = form.cleaned_data.get('email')
            username = first_name.strip().lower() + '_' + last_name.strip().lower()

            if User.objects.filter(username=username).exists():
                messages.error(request, f'{username} is already registered. Please choose a different username.')
            else:
                user = User.objects.create_user(
                    username=username, 
                    first_name=first_name, 
                    last_name=last_name,
                    email=email,
                    password='@@aa12345',
                )
                if user:
                    name = first_name + ' ' + last_name
                    image = form.cleaned_data['image']
                    grade = form.cleaned_data.get('grade')
                    age = form.cleaned_data.get('age')
                    gender = form.cleaned_data.get('gender')
                    parent_phone = form.cleaned_data.get('parent_phone')
                    student = Student(
                        user=user, 
                        name=name, 
                        image=image,
                        grade=grade, 
                        gender=gender,
                        age=age,
                        parent_phone=parent_phone,
                    )
                    if student:
                        student.save()
                        messages.success(request, f'{name} is registered successfully!')
                        return redirect('base:dashboard')
                    else:
                        user.delete()
                        messages.error(request, 'Someting went wrong while registering Student')
                        
                return redirect('base:student_register')
        else:
            messages.error(request, 'Someting went wrong while registering Student')
    else:
        form = StudentCreationForm()

    context = {
        'form': form,
        'type': 'student',
    }
    return render(request, 'dashboard/register.html', context)


@login_required
@user_passes_test(lambda user: user.is_staff)
def parent_register(request):
    if request.method == 'POST':
        form = ParentCreationForm(request.POST)
        if form.is_valid():
            first_name = form.cleaned_data.get('first_name')
            last_name = form.cleaned_data.get('last_name')
            email = form.cleaned_data.get('email')
            username = first_name.strip().lower() + '_' + last_name.strip().lower()

            if User.objects.filter(username=username).exists():
                messages.error(request, f'{username} is already registered. Please choose a different username.')
            else:
                user = User.objects.create_user(
                    username=username, 
                    first_name=first_name, 
                    last_name=last_name,
                    email=email,
                    password='@@aa12345',
                )
                if user:
                    name = first_name + ' ' + last_name
                    phone = form.cleaned_data.get('phone')
                    parent = Parent(
                        user=user, 
                        name=name,
                        phone=phone,
                    )
                    if parent:
                        parent.save()
                        messages.success(request, f'{name} is registered successfully!')
                        return redirect('base:dashboard')
                    else:
                        user.delete()
                        messages.error(request, 'Someting went wrong while registering parent')

                return redirect('base:parent_register')
            
    else:
        form = ParentCreationForm()

    context = {
        'form': form,
        'type': 'parent',
    }
    return render(request, 'dashboard/register.html', context)


@login_required
@user_passes_test(lambda user: user.is_staff)
def staff_register(request):
    if request.method == 'POST':
        form = StaffCreationForm(request.POST)
        if form.is_valid():
            first_name = form.cleaned_data.get('first_name')
            last_name = form.cleaned_data.get('last_name')
            email = form.cleaned_data.get('email')
            username = first_name.strip().lower() + '_' + last_name.strip().lower()

            if User.objects.filter(username=username).exists():
                messages.error(request, f'{username} is already registered. Please choose a different username.')
            else:
                user = User.objects.create_user(
                    username=username, 
                    first_name=first_name, 
                    last_name=last_name,
                    email=email,
                    password='@@aa12345',
                    is_staff=True,
                )
                if user:
                    name = first_name + ' ' + last_name
                    gender = form.cleaned_data.get('gender')
                    phone = form.cleaned_data.get('phone')
                    staff = Staff(
                        user=user, 
                        name=name,
                        gender=gender,
                        phone=phone,
                    )
                    if staff:
                        staff.save()
                        messages.success(request, f'{name} is registered successfully!')
                        return redirect('base:dashboard')
                    else:
                        user.delete()
                        messages.error(request, 'Someting went wrong while registering staff')

                return redirect('base:staff_register')
            
    else:
        form = StaffCreationForm()

    context = {
        'form': form,
        'type': 'staff',
    }
    return render(request, 'dashboard/register.html', context)


@login_required
@user_passes_test(lambda user: user.is_staff)
def student_register_with_file(request):
    if request.method == 'POST':
        form = ExcelImportForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data['file']
            if not file.name.endswith('.xlsx'):
                messages.error(request, 'Wrong file type. Please make sure the file type is Excel (i.e Filename ends with .xlsx)!')
                return render(request, 'dashboard/student_register_with_file.html')

            wb = load_workbook(file)
            sheet = wb.active

            complete = []
            success_count = 0
            count = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):
                count += 1
                try:
                    student_data = {
                        'name': str(row[0]),
                        'grade': str(row[1]),
                        'gender': str(row[2]),
                        'age': str(row[3]),
                    }
                except:
                    messages.error(request, 'It seems like the file is not supplying the required data to register Teachers.')
                    break
                
                student_form = StudentCreationFormWithFile(student_data)
                if student_form.is_valid():
                    name = student_form.cleaned_data['name']
                    first_name, last_name = student_form.cleaned_data['name'].split()[:2]
                    username = first_name.lower() + '_' + last_name.lower()
                    
                    if User.objects.filter(username=username).exists():
                        messages.error(request, f'{username} is already registered. Please choose a different username.')
                        continue
                    else:
                        user = User.objects.create_user(
                            username=username, 
                            first_name=first_name, 
                            last_name=last_name,
                            password='@@aa12345',
                        )
                        if user:
                            name = first_name + ' ' + last_name
                            grade = student_form.cleaned_data.get('grade')
                            age = student_form.cleaned_data.get('age')
                            gender = student_form.cleaned_data.get('gender')
                            student = Student(
                                user=user, 
                                name=name, 
                                grade=grade, 
                                gender=gender,
                                age=age,
                            )
                            if student:
                                student.save()
                                complete.append(student.student_id)
                                success_count += 1
                                continue
                            else:
                                user.delete()
                    messages.error(request, f'Something went wrong while registering {name}!')
                    count += 1
                    continue
                                
                else:
                    messages.error(request, f'Make sure the fields for {row[0]} are all valid!')
            context = {
                'students': Student.objects.order_by('name'),
                'complete': complete,
                'success_count': success_count,
                'count': count,
            }
            return render(request, 'dashboard/student_file_register_complete.html', context)
        else:
            messages.error(request, 'Make sure your file is excel type')
    return render(request, 'dashboard/student_register_with_file.html')


@login_required
@user_passes_test(lambda user: user.is_staff)
def teacher_register_with_file(request):
    if request.method == 'POST':
        form = ExcelImportForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data['file']
            if not file.name.endswith('.xlsx'):
                messages.error(request, 'Wrong file type. Please make sure the file type is Excel (i.e Filename ends with .xlsx)!')
                return render(request, 'dashboard/teacher_register_with_file.html')

            wb = load_workbook(file)
            sheet = wb.active

            complete = []
            success_count = 0
            count = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):
                count += 1
                try:
                    teacher_data = {
                        'name': str(row[0]),
                        'gender': str(row[1]),
                    }
                except:
                    messages.error(request, 'It seems like the file is not supplying the required data to register Teachers.')
                    break

                teacher_form = TeacherCreationFormWithFile(teacher_data)
                if teacher_form.is_valid():
                    name = teacher_form.cleaned_data['name']
                    first_name, last_name = teacher_form.cleaned_data['name'].split()[:2]
                    username = first_name.lower() + '_' + last_name.lower()
                    
                    if User.objects.filter(username=username).exists():
                        messages.error(request, f'{username} is already registered.')
                        continue
                    else:
                        user = User.objects.create_user(
                            username=username, 
                            first_name=first_name, 
                            last_name=last_name,
                            password='@@aa12345',
                        )
                        if user:
                            name = first_name + ' ' + last_name
                            gender = teacher_form.cleaned_data.get('gender')
                            teacher = Teacher(
                                user=user, 
                                name=name,
                                gender=gender,
                            )
                            if teacher:
                                teacher.save()
                                complete.append(teacher.teacher_id)
                                success_count += 1
                                continue
                            else:
                                user.delete()
                    messages.error(request, f'Something went wrong while registering {name}!')
                    count += 1
                    continue
                                
                else:
                    messages.error(request, f'Make sure the fields for {row[0]} are all valid!')
            context = {
                'teachers': Teacher.objects.order_by('name'),
                'complete': complete,
                'success_count': success_count,
                'count': count,
            }
            return render(request, 'dashboard/teacher_file_register_complete.html', context)
        else:
            messages.error(request, 'Make sure your file is excel type')
    return render(request, 'dashboard/teacher_register_with_file.html')




@login_required
@user_passes_test(lambda user: user.is_staff)
def student_dashboard(request, grade='all'):
    students = Student.objects.filter(grade=grade)
    if grade == 'all':
        students = Student.objects.all()
    students = students.order_by('name')
    context = {
        'students': students,
    }
    return render(request, 'dashboard/student_dashboard.html', context)


@login_required
@user_passes_test(lambda user: user.is_staff)
def teacher_dashboard(request):
    teachers = Teacher.objects.order_by('name')
    context = {
        'teachers': teachers,
    }
    return render(request, 'dashboard/teacher_dashboard.html', context)


@login_required
@user_passes_test(lambda user: user.is_staff)
def staff_dashboard(request):
    staffs = Staff.objects.order_by('name')
    context = {
        'staffs': staffs,
    }
    return render(request, 'dashboard/staff_dashboard.html', context)

@login_required
@user_passes_test(lambda user: user.is_staff)
def parent_dashboard(request):
    parents = Parent.objects.order_by('name')
    context = {
        'parents': parents,
    }
    return render(request, 'dashboard/parent_dashboard.html', context)



@login_required
@user_passes_test(lambda user: user.is_staff)
def student_update(request, pk):
    student = get_object_or_404(Student, student_id=pk)
    user_update = student.user
    if request.method == 'POST':
        form = StudentCreationForm(request.POST, request.FILES)
        if form.is_valid():
            user_update.first_name = form.cleaned_data['first_name']
            user_update.last_name = form.cleaned_data['last_name']
            user_update.email = form.cleaned_data['email']
            user_update.username = user_update.first_name.strip().lower() + '_' + user_update.last_name.strip().lower()
             
            user_update.save()
            if user_update:
                student.name = user_update.first_name + ' ' + user_update.last_name
                student.image = form.cleaned_data['image']
                student.grade = form.cleaned_data['grade']
                student.age = form.cleaned_data['age']
                student.gender = form.cleaned_data['gender']
                student.parent_phone = form.cleaned_data['parent_phone']
                if student.image:
                    student.image = form.cleaned_data['image']
                student.save()
                if student:
                    messages.success(request, 'Student is updated successfully!')
                    return redirect('base:student_dashboard')
                else:
                    messages.error(request, 'Error occured while updating student!')
                    return redirect('base:student_update')

            
    else:
        form = StudentCreationForm(instance=student)
    
    exclude = ['first_name', 'last_name', 'email']
    context = {
        'form': form,
        'type': 'student',
        'id': student.student_id,
        'user_update': user_update,
        'exclude': exclude,
    }

    return render(request, 'dashboard/update.html', context)


@login_required
@user_passes_test(lambda user: user.is_staff)
def teacher_update(request, pk):
    teacher = get_object_or_404(Teacher, teacher_id=pk)
    user_update = teacher.user
    if request.method == 'POST':
        form = TeacherCreationForm(request.POST, instance=teacher)
        if form.is_valid():
            user_update.first_name = form.cleaned_data['first_name']
            user_update.last_name = form.cleaned_data['last_name']
            user_update.email = form.cleaned_data['email']
            user_update.username = user_update.first_name.strip().lower() + '_' + user_update.last_name.strip().lower()
             
            user_update.save()
            if user_update:
                teacher.name = user_update.first_name + ' ' + user_update.last_name
                teacher.gender = form.cleaned_data['gender']
                teacher.department = form.cleaned_data['department']
                teacher.experience = form.cleaned_data['experience']
                teacher.phone = form.cleaned_data['phone']
                teacher.save()
                if teacher:
                    messages.success(request, 'Teacher is updated successfully!')
                    return redirect('base:teacher_dashboard')
                else:
                    messages.error(request, 'Error occured while updating teacher!')
                    return redirect('base:teacher_update')

            
    else:
        form = TeacherCreationForm(instance=teacher)
    
    exclude = ['first_name', 'last_name', 'email']
    context = {
        'form': form,
        'type': 'teacher',
        'id': teacher.teacher_id,
        'user_update': user_update,
        'exclude': exclude,
    }

    return render(request, 'dashboard/update.html', context)



@login_required
@user_passes_test(lambda user: user.is_staff)
def parent_update(request, pk):
    parent = get_object_or_404(Parent, parent_id=pk)
    user_update = parent.user
    if request.method == 'POST':
        form = ParentCreationForm(request.POST, instance=parent)
        if form.is_valid():
            user_update.first_name = form.cleaned_data['first_name']
            user_update.last_name = form.cleaned_data['last_name']
            user_update.email = form.cleaned_data['email']
            user_update.username = user_update.first_name.strip().lower() + '_' + user_update.last_name.strip().lower()
             
            user_update.save()
            if user_update:
                parent.name = user_update.first_name + ' ' + user_update.last_name
                parent.phone = form.cleaned_data['phone']
                parent.save()
                if parent:
                    messages.success(request, 'Parent is updated successfully!')
                    return redirect('base:parent_dashboard')
                else:
                    messages.error(request, 'Error occured while updating Parent!')
                    return redirect('base:parent_update')

            
    else:
        form = ParentCreationForm(instance=parent)
    
    exclude = ['first_name', 'last_name', 'email']
    context = {
        'form': form,
        'type': 'parent',
        'id': parent.parent_id,
        'user_update': user_update,
        'exclude': exclude,
    }

    return render(request, 'dashboard/update.html', context)


@login_required
@user_passes_test(lambda user: user.is_staff)
def staff_update(request, pk):
    staff = get_object_or_404(Staff, staff_id=pk)
    user_update = staff.user
    if request.method == 'POST':
        form = StaffCreationForm(request.POST, instance=staff)
        if form.is_valid():
            user_update.first_name = form.cleaned_data['first_name']
            user_update.last_name = form.cleaned_data['last_name']
            user_update.email = form.cleaned_data['email']
            user_update.username = user_update.first_name.strip().lower() + '_' + user_update.last_name.strip().lower()
             
            user_update.save()
            if user_update:
                staff.name = user_update.first_name + ' ' + user_update.last_name
                staff.gender = form.cleaned_data['gender']
                staff.phone = form.cleaned_data['phone']
                staff.save()
                if staff:
                    messages.success(request, 'staff is updated successfully!')
                    return redirect('base:staff_dashboard')
                else:
                    messages.error(request, 'Error occured while updating staff!')
                    return redirect('base:staff_update')

            
    else:
        form = StaffCreationForm(instance=staff)
    
    exclude = ['first_name', 'last_name', 'email']
    context = {
        'form': form,
        'type': 'staff',
        'id': staff.staff_id,
        'user_update': user_update,
        'exclude': exclude,
    }

    return render(request, 'dashboard/update.html', context)



# Profile section
@login_required
def student_profile(request, pk):
    student = Student.objects.get(student_id=pk)
    if request.user != student.user and not request.user.is_staff:
        return HttpResponse("You can't access this profile!")
    scores = student.scores.all()
    context = {
        'student': student,
        'scores': scores,
    }
    return render(request, 'profile/student_profile.html', context)

@login_required
def change_password(request):
    if request.method == 'POST':
        current_password = request.POST['current_password']
        new_password = request.POST['new_password']
        confirm_password = request.POST['confirm_password']

        user = authenticate(username=request.user.username, password=current_password)
        if user:
            if new_password == confirm_password:
                user.set_password(new_password)
                user.save()
                login(request, user)
                messages.success(request, 'Password is successfully!')
                return redirect('base:home')
            else:
                messages.error(request, 'The New Password and Confirm Password do not match!')
                return redirect('base:change_password')
        else:
            messages.error(request, 'Wrong Password. Please try again!')
            return redirect('base:change_password')

    return render(request, 'dashboard/change_password.html')




@login_required
@user_passes_test(lambda user: user.is_staff)
def result_dashboard(request, grade=None):
    if grade == 'all':
        students = Student.objects.order_by('name')
    else:
        students = Student.objects.filter(grade=grade)
    
    context = {
        'students': students,
    }
    return render(request, 'dashboard/result_dashboard.html', context)



@login_required
@user_passes_test(lambda user: user.is_staff)
def publish_result(request):
    if request.method == 'POST':
        form = ScoreForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data['file']
            if not file.name.endswith('.xlsx'):
                messages.error(request, 'Wrong file type. Please make sure the file type is Excel (i.e Filename ends with .xlsx)!')
                return render(request, 'dashboard/publish_result.html')

            wb = load_workbook(file)
            sheet = wb.active

            semester = form.cleaned_data['semester']
            grade = form.cleaned_data['grade']

            if 1 <= int(grade) <= 5:
                return publish_for_grade_1_to_5(request, sheet, semester, grade)  # util function that publishes resutl for grade 5 students
            
            elif 5 < int(grade) <= 8:
                return publish_for_grade_6_to_8(request, sheet, semester, grade)  # util function that publishes resutl for grade 5 students
            
            elif 8 < int(grade) <= 10:
                return publish_for_grade_9_and_10(request, sheet, semester, grade)  # util function that publishes resutl for grade 5 students
        else:
            messages.error(request, 'Make sure your file is excel type')
    
    form = ScoreForm()
    context = {
        'form': form,
    }
    return render(request, 'dashboard/publish_result.html', context)


# restriction redirect veiw
def restricted_view(request):
    return render(request, 'base/restricted.html')